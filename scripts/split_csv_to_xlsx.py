"""
IMI CSV→xlsx 拆分脚本
将 {case}_trajs.csv + {case}_imi.csv 拆分为 {case}_traj1/2/3.xlsx

用法: python split_csv_to_xlsx.py <case_name> <r1_start>,<r1_end> <r2_start>,<r2_end> <r3_start>,<r3_end>
示例: python split_csv_to_xlsx.py 225CKH 1,15 16,30 31,45
"""
import sys
import csv
import openpyxl


def read_trajs_csv(path):
    """读取轨迹CSV，返回3条轨迹各6行的列表"""
    with open(path, newline='') as f:
        rows = list(csv.reader(f))
    return [rows[i:i + 6] for i in range(0, len(rows), 6)]


def read_imi_csv(path):
    """读取IMI CSV，返回 {label: [values...]} 字典（labels: depth, angle, exit）"""
    with open(path, newline='') as f:
        rows = list(csv.reader(f))
    return {row[0].strip(): row[1:] for row in rows if row[0].strip()}


def make_single_xlsx(traj_data, imi_data, id_start, id_end, output_path):
    """生成一个xlsx文件"""
    wb = openpyxl.Workbook()

    # --- traj sheet ---
    ws_traj = wb.active
    ws_traj.title = 'traj'
    for r, row_data in enumerate(traj_data, 1):
        ws_traj.cell(row=r, column=1, value=row_data[0])
        ws_traj.cell(row=r, column=2, value=row_data[1])
        val = row_data[2]
        ws_traj.cell(row=r, column=3, value=float(val) if '.' in str(val) else int(val))

    # --- deposits sheet ---
    ws_dep = wb.create_sheet('deposits')
    start = id_start - 1  # 0-indexed
    end = id_end
    n = end - start

    ws_dep.cell(row=1, column=1, value='deposit#')
    for i in range(n):
        ws_dep.cell(row=1, column=i + 2, value=i + 1)

    for r_offset, key in enumerate(['depth', 'angle', 'exit']):
        ws_dep.cell(row=r_offset + 2, column=1, value=key)
        vals = imi_data[key][start:end]
        for i, v in enumerate(vals):
            ws_dep.cell(row=r_offset + 2, column=i + 2, value=float(v) if '.' in v else int(v))

    for r_offset, label in enumerate(['R', 'A', 'S']):
        ws_dep.cell(row=r_offset + 5, column=1, value=label)

    wb.save(output_path)
    return n


def main():
    if len(sys.argv) != 5:
        print("用法: python split_csv_to_xlsx.py <case_name> <r1_start>,<r1_end> <r2_start>,<r2_end> <r3_start>,<r3_end>")
        print("示例: python split_csv_to_xlsx.py 225CKH 1,15 16,30 31,45")
        sys.exit(1)

    case = sys.argv[1]
    ranges = []
    for i in range(2, 5):
        parts = sys.argv[i].split(',')
        ranges.append((int(parts[0]), int(parts[1])))

    trajs = read_trajs_csv(f'{case}_trajs.csv')
    imi = read_imi_csv(f'{case}_imi.csv')

    for idx, ((start, end), traj) in enumerate(zip(ranges, trajs)):
        traj_num = idx + 1
        fname = f'{case}_traj{traj_num}.xlsx'
        n = make_single_xlsx(traj, imi, start, end, fname)
        print(f'{fname}: traj{traj_num}, deposits {start}-{end} ({n}个)')

    print('完成')


if __name__ == '__main__':
    main()
